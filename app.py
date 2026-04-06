import os
from flask import Flask, request, redirect, render_template_string, send_file, session
from openpyxl import Workbook
from sqlalchemy import create_engine, Column, Integer, String, Float
from sqlalchemy.orm import declarative_base, sessionmaker
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime, timedelta
import smtplib
from email.message import EmailMessage

# =========================
# APP & DB SETUP
# =========================
app = Flask(__name__)
app.secret_key = "secret123"

Base = declarative_base()
engine = create_engine("sqlite:///data.db")
SessionDB = sessionmaker(bind=engine)
db = SessionDB()

# =========================
# USERS
# =========================
users = {"admin": {"password": "admin", "role": "admin"}}

# =========================
# DB MODELY
# =========================
class Zakazka(Base):
    __tablename__ = "zakazky"
    id = Column(Integer, primary_key=True)
    nazev = Column(String)
    owner = Column(String)


class Polozka(Base):
    __tablename__ = "polozky"
    id = Column(Integer, primary_key=True)
    zakazka_id = Column(Integer)
    nazev = Column(String)
    material = Column(String)
    cena_materialu = Column(Float)
    hodiny = Column(Float)
    sazba = Column(Float)
    datum = Column(String)


Base.metadata.create_all(engine)

# =========================
# HELPERS
# =========================
def current_user():
    return session.get("user")


def is_admin():
    return users.get(current_user(), {}).get("role") == "admin"

# =========================
# EMAIL SETTINGS
# =========================
EMAIL_FROM = "tvuj.email@gmail.com"
EMAIL_PASSWORD = "tvuj_app_password"
EMAIL_DAILY = "pecenyjirik@gmail.com"
EMAIL_WEEKLY = "pecenyjiri@gmail.com"

def send_email(to_email, subject, body, attachment_path):
    msg = EmailMessage()
    msg["From"] = EMAIL_FROM
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)

    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            data = f.read()
            name = os.path.basename(attachment_path)
            msg.add_attachment(data, maintype="application", subtype="octet-stream", filename=name)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_FROM, EMAIL_PASSWORD)
        smtp.send_message(msg)

# =========================
# EXCEL EXPORT
# =========================
def export_to_excel(zakazka_id):
    wb = Workbook()
    ws = wb.active

    zakazka = db.query(Zakazka).get(zakazka_id)

    ws.append([f"Zakázka: {zakazka.nazev}"])
    ws.append([])
    ws.append(["Název", "Materiál", "Cena materiálu", "Hodiny", "Sazba", "Datum", "Cena"])

    polozky = db.query(Polozka).filter_by(zakazka_id=zakazka_id).all()
    total = 0
    for p in polozky:
        cena = p.cena_materialu + (p.hodiny * p.sazba)
        total += cena
        ws.append([p.nazev, p.material, p.cena_materialu, p.hodiny, p.sazba, p.datum, cena])

    ws.append([])
    ws.append(["", "", "", "", "", "Celkem", total])
    filename = f"{zakazka.nazev}.xlsx"
    wb.save(filename)
    return filename

# =========================
# HTML TEMPLATES
# =========================
BOOTSTRAP = """
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
"""

LOGIN_HTML = BOOTSTRAP + """
<div class="container mt-5">
<h2>Přihlášení</h2>
<form method='post'>
<input name='user' class="form-control mb-2" placeholder='Uživatel'>
<input name='password' type='password' class="form-control mb-2" placeholder='Heslo'>
<button class="btn btn-primary">Přihlásit</button>
</form>
</div>
"""

INDEX_HTML = BOOTSTRAP + """
<div class="container mt-4">
<h1 class="mb-4">Zakázky ({{user}})</h1>

<a href='/logout' class="btn btn-danger mb-3">Odhlásit</a>

<form method='post' action='/nova' class="mb-4">
  <div class="input-group">
    <input name='nazev' class="form-control" placeholder='Nová zakázka'>
    <button class="btn btn-primary">Vytvořit</button>
  </div>
</form>

<div class="list-group">
{% for z in zakazky %}
  <div class="list-group-item d-flex justify-content-between align-items-center">
    <a href='/zakazka/{{z.id}}'>
      {{z.nazev}} <small class="text-muted">({{z.owner}})</small>
    </a>
    <form method='post' action='/edit_zakazka/{{z.id}}' class="d-flex">
      <input name='nazev' value='{{z.nazev}}' class="form-control form-control-sm me-2">
      <button class="btn btn-sm btn-warning">Upravit</button>
    </form>
  </div>
{% endfor %}
</div>
</div>
"""

DETAIL_HTML = BOOTSTRAP + """
<div class="container mt-4">

<h2 class="mb-3">{{z.nazev}}</h2>

<a href='/' class="btn btn-secondary mb-3">Zpět</a>

<form method='post' action='/pridat/{{z.id}}' class="card p-3 mb-4">
  <div class="row g-2">
    <div class="col"><input name='nazev' class="form-control" placeholder='Název'></div>
    <div class="col"><input name='material' class="form-control" placeholder='Materiál'></div>
    <div class="col"><input name='cena_materialu' class="form-control" placeholder='Cena'></div>
    <div class="col"><input name='hodiny' class="form-control" placeholder='Hodiny'></div>
    <div class="col"><input name='sazba' class="form-control" placeholder='Sazba'></div>
    <div class="col"><input name='datum' type='date' class="form-control"></div>
  </div>
  <button class="btn btn-success mt-2">Přidat</button>
</form>

<table class="table table-striped">
<tr>
<th>Název</th><th>Materiál</th><th>Cena mat.</th><th>Hodiny</th><th>Sazba</th><th>Datum</th><th>Akce</th>
</tr>

{% for p in polozky %}
<tr>
<form method='post' action='/edit_polozka/{{p.id}}'>
<td><input name='nazev' class="form-control form-control-sm" value='{{p.nazev}}'></td>
<td><input name='material' class="form-control form-control-sm" value='{{p.material}}'></td>
<td><input name='cena_materialu' class="form-control form-control-sm" value='{{p.cena_materialu}}'></td>
<td><input name='hodiny' class="form-control form-control-sm" value='{{p.hodiny}}'></td>
<td><input name='sazba' class="form-control form-control-sm" value='{{p.sazba}}'></td>
<td><input name='datum' type='date' class="form-control form-control-sm" value='{{p.datum}}'></td>
<td>
  <button class="btn btn-sm btn-warning mb-1">Upravit</button>
  <a href='/delete_polozka/{{p.id}}' class="btn btn-sm btn-danger mb-1">Smazat</a>
</td>
</form>
</tr>
{% endfor %}
</table>

<a href='/export/{{z.id}}' class="btn btn-outline-primary">Export do Excelu</a>

</div>
"""

# =========================
# ROUTES
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = request.form["user"]
        p = request.form["password"]
        if u in users and users[u]["password"] == p:
            session["user"] = u
            return redirect("/")
    return render_template_string(LOGIN_HTML)

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/")
def index():
    if not current_user():
        return redirect("/login")
    zakazky = db.query(Zakazka).all() if is_admin() else db.query(Zakazka).filter_by(owner=current_user()).all()
    return render_template_string(INDEX_HTML, zakazky=zakazky, user=current_user())

@app.route("/zakazka/<int:id>")
def detail(id):
    z = db.query(Zakazka).get(id)
    polozky = db.query(Polozka).filter_by(zakazka_id=id).all()
    if not is_admin() and z.owner != current_user(): return "Access denied"
    return render_template_string(DETAIL_HTML, z=z, polozky=polozky)

@app.route("/nova", methods=["POST"])
def nova():
    z = Zakazka(nazev=request.form["nazev"], owner=current_user())
    db.add(z)
    db.commit()
    return redirect("/")

@app.route("/edit_zakazka/<int:id>", methods=["POST"])
def edit_zakazka(id):
    z = db.query(Zakazka).get(id)
    if not z: return "Zakázka nenalezena"
    if not is_admin() and z.owner != current_user(): return "Access denied"
    new_name = request.form.get("nazev")
    if new_name: z.nazev = new_name; db.commit()
    return redirect("/")

@app.route("/pridat/<int:id>", methods=["POST"])
def pridat(id):
    try:
        p = Polozka(
            zakazka_id=id,
            nazev=request.form["nazev"],
            material=request.form["material"],
            cena_materialu=float(request.form["cena_materialu"].replace(",", ".")),
            hodiny=float(request.form["hodiny"].replace(",", ".")),
            sazba=float(request.form["sazba"].replace(",", ".")),
            datum=request.form["datum"]
        )
        db.add(p)
        db.commit()
    except ValueError:
        return "Chyba: čísla nejsou správně"
    return redirect(f"/zakazka/{id}")

@app.route("/edit_polozka/<int:id>", methods=["POST"])
def edit_polozka(id):
    p = db.query(Polozka).get(id)
    if not p: return "Položka nenalezena"
    zak = db.query(Zakazka).get(p.zakazka_id)
    if not is_admin() and zak.owner != current_user(): return "Access denied"
    try:
        p.nazev = request.form.get("nazev")
        p.material = request.form.get("material")
        p.cena_materialu = float(request.form.get("cena_materialu").replace(",", "."))
        p.hodiny = float(request.form.get("hodiny").replace(",", "."))
        p.sazba = float(request.form.get("sazba").replace(",", "."))
        p.datum = request.form.get("datum")
        db.commit()
    except ValueError:
        return "Chyba: čísla nejsou správně"
    return redirect(f"/zakazka/{p.zakazka_id}")

@app.route("/delete_polozka/<int:id>")
def delete_polozka(id):
    p = db.query(Polozka).get(id)
    if not p: return "Položka nenalezena"
    zak = db.query(Zakazka).get(p.zakazka_id)
    if not is_admin() and zak.owner != current_user(): return "Access denied"
    db.delete(p); db.commit()
    return redirect(f"/zakazka/{zak.id}")

@app.route("/export/<int:id>")
def export(id):
    filename = export_to_excel(id)
    return send_file(filename, as_attachment=True)

# =========================
# AUTOMATICKÉ ZÁLOHY
# =========================
def daily_backup():
    print("Spouštím denní zálohu:", datetime.now())
    yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    filename = f"daily_backup_{yesterday}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Zakázka", "Název", "Materiál", "Cena materiálu", "Hodiny", "Sazba", "Datum", "Celkem"])
    polozky = db.query(Polozka).filter(Polozka.datum >= yesterday).all()
    for p in polozky:
        zak = db.query(Zakazka).get(p.zakazka_id)
        cena = p.cena_materialu + (p.hodiny * p.sazba)
        ws.append([zak.nazev, p.nazev, p.material, p.cena_materialu, p.hodiny, p.sazba, p.datum, cena])
    wb.save(filename)
    send_email(EMAIL_DAILY, "Denní záloha zakázek", "Přiložen denní export nových položek.", filename)
    os.remove(filename)

def weekly_backup():
    print("Spouštím týdenní zálohu:", datetime.now())
    filename = f"weekly_backup_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Zakázka", "Název", "Materiál", "Cena materiálu", "Hodiny", "Sazba", "Datum", "Celkem"])
    zakazky = db.query(Zakazka).all()
    for z in zakazky:
        polozky = db.query(Polozka).filter_by(zakazka_id=z.id).all()
        for p in polozky:
            cena = p.cena_materialu + (p.hodiny * p.sazba)
            ws.append([z.nazev, p.nazev, p.material, p.cena_materialu, p.hodiny, p.sazba, p.datum, cena])
    wb.save(filename)
    send_email(EMAIL_WEEKLY, "Týdenní záloha zakázek", "Přiložen týdenní export všech položek.", filename)
    os.remove(filename)

scheduler = BackgroundScheduler()
scheduler.add_job(daily_backup, "cron", hour=0, minute=0)
scheduler.add_job(weekly_backup, "cron", day_of_week="sun", hour=1, minute=0)
scheduler.start()

# =========================
# START APP
# =========================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
